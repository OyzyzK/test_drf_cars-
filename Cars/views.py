import csv
import openpyxl
from django.http import HttpResponse
from rest_framework import viewsets
from openpyxl.utils import get_column_letter
from rest_framework import generics
from .serializers import CarItemSerializer, ProducerItemSerializer, FullCountrySerializer, FullProducerSerializer, \
    FullCarSerialzer, CommentItemSerializer
from .models import CountryItem, ProducerItem, CarItem, CommentItem
from .permissions import AdminPermission, CommentPermission


def down_file_countries(request, **kwargs):
    queryset = CountryItem.objects.all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Countries"
    row_num = 0
    columns = [
        ("ID", 5),
    ]
    for col_num in range(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][0]
        ws.column_dimensions[get_column_letter(col_num + 1)].width = columns[col_num][1]

    for obj in queryset:
        row_num += 1
        row = [
            obj.pk,
            obj.name_country
        ]
        for col_num in range(len(row)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = row[col_num]
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=test.xlsx'
    wb.save(response)
    return response


def exportcsv_countries(request):
    countries = CountryItem.objects.all()
    response = HttpResponse('text/csv')
    response['Content-Disposition'] = 'attachment; filename=countries.csv'
    writer = csv.writer(response)
    writer.writerow(['ID', 'Name'])
    country = countries.values_list('id', 'name_country')
    for ctr in country:
        writer.writerow(ctr)
    return response


class CountriesViewSet(viewsets.ModelViewSet):
    serializer_class = FullCountrySerializer
    permission_classes = [AdminPermission]

    def get_queryset(self):
        return CountryItem.objects.all()


def down_file_producers(request, **kwargs):
    queryset = ProducerItem.objects.all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Producers"
    row_num = 0
    columns = [
        ("ID", 5),
    ]
    for col_num in range(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][0]
        ws.column_dimensions[get_column_letter(col_num + 1)].width = columns[col_num][1]

    for obj in queryset:
        row_num += 1
        row = [
            obj.pk,
            obj.name_producer,
            str(obj.name_country)
        ]
        for col_num in range(len(row)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = row[col_num]
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=test.xlsx'
    wb.save(response)
    return response


def exportcsv_producers(request):
    producers = ProducerItem.objects.all()
    response = HttpResponse('text/csv')
    response['Content-Disposition'] = 'attachment; filename=producers.csv'
    writer = csv.writer(response)
    writer.writerow(['ID', 'Name_prod', 'Name_country'])
    producer = producers.values_list('id', 'name_producer', 'name_country_id', 'name_country')
    for prd in producer:
        writer.writerow(prd)
    return response


class ProducerViewSet(viewsets.ModelViewSet):
    serializer_class = FullProducerSerializer
    permission_classes = [AdminPermission]

    def get_queryset(self):
        return ProducerItem.objects.all()


def down_file_cars(request, **kwargs):
    queryset = CarItem.objects.all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cars"
    row_num = 0
    columns = [
        ("ID", 5),
    ]
    for col_num in range(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][0]
        ws.column_dimensions[get_column_letter(col_num + 1)].width = columns[col_num][1]

    for obj in queryset:
        row_num += 1
        row = [
            obj.pk,
            obj.name_car,
            str(obj.name_producer),
            str(obj.year_start),
            str(obj.year_end)
        ]
        for col_num in range(len(row)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = row[col_num]
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=cars.xlsx'
    wb.save(response)
    return response


def exportcsv_cars(request):
    cars = CarItem.objects.all()
    response = HttpResponse('text/csv')
    response['Content-Disposition'] = 'attachment; filename=cars.csv'
    writer = csv.writer(response)
    writer.writerow(['ID', 'Name_Car', 'Name_Prod', 'Year_Start', 'Year_End'])
    car = cars.values_list('id', 'name_car', 'name_producer_id', 'name_producer', 'year_start', 'year_end')
    for cr in car:
        writer.writerow(cr)
    return response


class CarsViewSet(viewsets.ModelViewSet):
    serializer_class = FullCarSerialzer
    permission_classes = [AdminPermission]

    def get_queryset(self):
        return CarItem.objects.all()


def down_file_comments(request, **kwargs):
    queryset = CommentItem.objects.all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Producers"
    row_num = 0
    columns = [
        ("ID", 5),
    ]
    for col_num in range(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][0]
        ws.column_dimensions[get_column_letter(col_num + 1)].width = columns[col_num][1]

    for obj in queryset:
        row_num += 1
        row = [
            obj.pk,
            obj.email_name,
            str(obj.name_car),
            str(obj.creation_date),
            obj.comment,

        ]
        for col_num in range(len(row)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = row[col_num]
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=test.xlsx'
    wb.save(response)
    return response


def exportcsv_comments(request):
    comments = CommentItem.objects.all()
    response = HttpResponse('text/csv')
    response['Content-Disposition'] = 'attachment; filename=students.csv'
    writer = csv.writer(response)
    writer.writerow(['ID', 'Name_prod', 'Name_country'])
    comment = comments.values_list('id', 'name_producer', 'name_country')
    for cmnt in comment:
        writer.writerow(cmnt)
    return response


class CommentViewSet(viewsets.ModelViewSet):
    serializer_class = CommentItemSerializer
    permission_classes = [CommentPermission]

    def get_queryset(self):
        return CommentItem.objects.all()




# Create your views here.
